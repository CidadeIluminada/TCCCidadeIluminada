# coding: UTF-8
from __future__ import absolute_import

import os
from datetime import datetime, date, timedelta
from StringIO import StringIO
from tempfile import mkstemp
from collections import Counter

from flask import request, abort, redirect, url_for, jsonify, send_file, flash, json
from flask.ext.admin import Admin, expose, AdminIndexView
from flask.ext.admin.actions import action
from flask.ext.admin.model import typefmt, InlineFormAdmin
from flask.ext.admin.contrib.fileadmin import FileAdmin
from flask.ext.admin.contrib.sqla import ModelView
from flask.ext.admin.contrib.sqla.fields import QuerySelectField
from flask.ext.admin.form.widgets import Select2Widget
from flask.ext.security import current_user, login_required, roles_accepted, roles_required
from sqlalchemy import text
from openpyxl import load_workbook
from xhtml2pdf import pisa
from wtforms.validators import Required

from cidadeiluminada.models import User, Role
from cidadeiluminada.protocolos import utils
from cidadeiluminada.protocolos.models import Regiao, Bairro, Logradouro, \
    Poste, ItemManutencao, Protocolo, OrdemServico, Servico, Equipamento, Material, \
    PrecoEquipamento
from cidadeiluminada.base import db

default_formatters = dict(typefmt.BASE_FORMATTERS, **{
    datetime: lambda view, value: utils.datetime_format(value),
    date: lambda view, value: utils.date_format(value),
})

_datetimepicker_locale = json.dumps({
    "applyLabel": "Aplicar",
    "cancelLabel": "Cancelar",
    "fromLabel": "De",
    "toLabel": "Até",
    "customRangeLabel": "Customizar",
    "daysOfWeek": [
        "Dom",
        "Seg",
        "Ter",
        "Qua",
        "Qui",
        "Sex",
        "Sab"
    ],
    "monthNames": [
        "Janeiro",
        "Fevereiro",
        "Março",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro"
    ],
    "firstDay": 0
})

form_widget_formats = {
    u'datetime': {
        'data-date-format': u'DD/MM/YYYY HH:mm:ss',
        'data-locale': _datetimepicker_locale,
        'data-separator': u' até '
    },
    u'date': {
        'data-date-format': u'DD/MM/YYYY',
        'data-locale': _datetimepicker_locale,
        'data-separator': u' até '
    }
}

form_args_formats = {
    u'datetime': {
        u'format': '%d/%m/%Y %H:%M:%S',
    },
    u'date': {
        'format': '%d/%m/%Y',
    }
}

preco_equipamento_labels = {
    'preco': u'Preço',
    'garantia': u'Dias de garantia',
    'inicio_vigencia': u'Início de vigência',
    'equipamento': u'Equipamento',
}


class IndexView(AdminIndexView):

    @expose('/')
    @login_required
    def index(self):
        if current_user.has_role('admin'):
            return redirect(url_for('.index_admin'))
        elif current_user.has_role('urbam'):
            return redirect(url_for('.index_urbam'))
        elif current_user.has_role('secretaria'):
            return redirect(url_for('.index_secretaria'))

    @expose('/admin')
    @roles_required('admin')
    def index_admin(self):
        return self.render('admin/index_admin.html')

    def _carregar_os(self):
        novas = OrdemServico.query.filter(OrdemServico.nova) \
            .order_by(OrdemServico.id.desc())
        em_servico = OrdemServico.query.filter(OrdemServico.em_servico) \
            .order_by(OrdemServico.id.desc())
        feitas = OrdemServico.query.filter(OrdemServico.feita) \
            .order_by(OrdemServico.id.desc())
        return {
            u'ordens_servico_novas': novas,
            u'ordens_servico_em_servico': em_servico,
            u'ordens_servico_feitas': feitas,
        }

    @expose('/secretaria')
    @roles_accepted('admin', 'secretaria')
    def index_secretaria(self):
        im_query = ItemManutencao.query.join(Poste).join(Logradouro).join(Bairro).join(Regiao) \
            .filter(ItemManutencao.status == 'aberto')
        regioes_select_map = {}
        regioes_qty_map = {}
        for regiao in db.session.query(Regiao.id, Regiao.nome):
            qty_im = im_query.filter(Regiao.id == regiao.id).count()
            regioes_select_map[regiao.id] = u'Região {}'.format(regiao.nome)
            regioes_qty_map[regiao.id] = qty_im
        ordens_servico = self._carregar_os()
        return self.render('admin/index_secretaria.html', regioes_map=regioes_select_map,
                           regioes_qty=regioes_qty_map, **ordens_servico)

    @expose('/urbam')
    @roles_accepted('admin', 'urbam')
    def index_urbam(self):
        ordens_servico = self._carregar_os()
        return self.render('admin/index_urbam.html', **ordens_servico)


class SecretariaAcessibleMixin():

    def is_accessible(self):
        has_secretaria = current_user.has_role('secretaria')
        return current_user.is_authenticated and has_secretaria


class _ModelView(SecretariaAcessibleMixin, ModelView):

    category = None

    column_type_formatters = default_formatters

    def __init__(self, *args, **kwargs):
        kwargs.setdefault('name', self.name)
        kwargs.setdefault('category', self.category)
        super(_ModelView, self).__init__(self.model, db.session, *args, **kwargs)


class _UserModelsView(_ModelView):
    def is_accessible(self):
        return current_user.has_role('admin')


class InlinePrecoEquipamentoForm(InlineFormAdmin):
    column_labels = preco_equipamento_labels


class EquipamentoView(_ModelView):
    model = Equipamento
    name = u'Equipamento'
    category = u'Equipamentos'

    column_labels = {
        'nome': u'Nome',
        u'precos': u'Preços',
        u'abreviacao': u'Abreviação',
    }

    form_excluded_columns = ['materiais']
    column_exclude_list = ['precos']

    inline_models = [InlinePrecoEquipamentoForm(PrecoEquipamento)]


class PrecoEquipamentoView(_ModelView):
    model = PrecoEquipamento
    name = u'Preços'
    category = u'Equipamentos'

    column_labels = preco_equipamento_labels

    form_args = {
        u'inicio_vigencia': form_args_formats[u'date'],
    }

    form_widget_args = {
        u'inicio_vigencia': form_widget_formats[u'date'],
    }


class RegiaoView(_ModelView):
    model = Regiao
    name = u'Região'
    category = u'Endereço'

    form_excluded_columns = ['bairros']

    @expose('/bairros')
    def get_bairros(self):
        regiao_id = request.args['regiao_id']
        regiao = self.model.query.get_or_404(regiao_id)
        im_query = ItemManutencao.query.join(Poste).join(Logradouro).join(Bairro) \
            .filter(ItemManutencao.status == 'aberto')
        bairros = []
        for bairro in regiao.bairros:
            qty_im = im_query.filter(Bairro.id == bairro.id).count()
            serialized = bairro.serialize()
            serialized['qty_im'] = qty_im
            bairros.append(serialized)
        return jsonify({
            'payload': bairros,
        })


class BairroView(_ModelView):
    model = Bairro
    name = 'Bairro'
    category = u'Endereço'

    form_excluded_columns = ['logradouros']

    column_labels = {
        'regiao': u'Região',
    }

    column_filters = ['nome', 'regiao']

    @expose('/itens_manutencao')
    def get_itens_manutencao(self):
        bairro_id = request.args['bairro_id']
        im_query = ItemManutencao.query.join(Poste).join(Logradouro).join(Bairro) \
            .filter(ItemManutencao.status == 'aberto').filter(Bairro.id == bairro_id)
        postes = []
        for item_manutencao in im_query:
            poste = item_manutencao.poste
            serialized = poste.serialize()

            serialized['label'] = u'{} - {} - Nº {}'.format(poste.logradouro.cep,
                                                             poste.logradouro.logradouro,
                                                             poste.numero)
            postes.append(serialized)
        return jsonify({
            'payload': postes,
        })


class LogradouroView(_ModelView):
    model = Logradouro
    name = 'Logradouro'
    category = u'Endereço'

    create_template = 'admin/model/edit_modelo_cep.html'
    edit_template = 'admin/model/edit_modelo_cep.html'

    form_columns = ('cep', 'bairro', 'logradouro')

    form_widget_args = {
        'logradouro': {
            'readonly': True,
        },
    }

    column_labels = {
        'cep': 'CEP',
    }

    column_filters = ['logradouro', 'cep', 'bairro', 'bairro.regiao']

    def _inject_bairros(self):
        bairros = Bairro.query
        bairro_id_nome = {bairro.nome: bairro.id for bairro in bairros}
        self._template_args['bairro_id_nome_map'] = bairro_id_nome

    @expose('/new/', methods=('GET', 'POST'))
    def create_view(self):
        self._inject_bairros()
        return super(LogradouroView, self).create_view()

    @expose('/edit/', methods=('GET', 'POST'))
    def edit_view(self):
        self._inject_bairros()
        return super(LogradouroView, self).edit_view()


class PosteView(_ModelView):
    def __init__(self, item_manutencao_view, *args, **kwargs):
        super(PosteView, self).__init__(*args, **kwargs)
        self.item_manutencao_view = item_manutencao_view

    model = Poste
    name = 'Postes'
    category = 'Protocolos'

    form_columns = ('logradouro', 'numero')

    edit_template = 'admin/model/edit_poste.html'

    column_labels = {
        'numero': u'Número',
    }

    @expose('/edit/', methods=('GET', 'POST'))
    def edit_view(self):
        cols = self.item_manutencao_view.get_list_columns()
        cols.remove(('poste', 'Poste'))
        self._template_args['list_columns'] = cols
        self._template_args['get_item_manutencao_value'] = self.item_manutencao_view.get_list_value
        return super(PosteView, self).edit_view()

    @expose(u'/logradouro/')
    def logradouro(self):
        logradouro_id = request.args.get(u'logradouro_id')
        postes_logradouro = self.model.query.filter_by(logradouro_id=logradouro_id).all()
        return jsonify({
            u'payload': {
                u'postes': postes_logradouro,
            }
        })

    @expose(u'/vincular_protocolo/', methods=[u'POST'])
    def vincular_protocolo(self):
        poste_id = int(request.form[u'poste_id'])
        cod_protocolo = request.form[u'cod_protocolo']
        criacao = datetime.strptime(request.form[u'criacao'], u'%Y-%m-%d')
        if poste_id:
            poste = Poste.query.get(poste_id)
        else:
            logradouro_id = int(request.form[u'logradouro_id'])
            numero = int(request.form[u'numero'])
            poste = Poste(logradouro_id=logradouro_id, numero=numero)
            db.session.add(poste)
        protocolo = Protocolo(poste=poste, cod_protocolo=cod_protocolo, criacao=criacao)
        db.session.add(protocolo)
        db.session.commit()
        return jsonify({
            u'payload': {
                u'protocolo': protocolo,
            }
        })
        pass


class ProtocoloView(_ModelView):
    model = Protocolo
    name = 'Protocolos 156'
    category = 'Protocolos'

    column_exclude_list = ['item_manutencao']
    column_filters = ['item_manutencao.id']

    named_filter_urls = True

    can_delete = False
    can_create = False

    form_excluded_columns = ['item_manutencao']
    edit_template = 'admin/model/edit_protocolo.html'

    column_labels = {
        'criacao': u'Criação',
        'cod_protocolo': u'Código do protocolo',
        'nome_municipe': u'Nome do munícipe',
        'contato_municipe': u'Contato do munícipe',
    }

    form_extra_fields = {
        'poste': QuerySelectField(query_factory=lambda: Poste.query.all(), allow_blank=True,
                                  widget=Select2Widget(),
                                  validators=[Required(u'Campo obrigatório')])
    }

    form_args = {
        u'criacao': form_args_formats[u'datetime'],
    }

    form_widget_args = {
        u'criacao': form_widget_formats[u'datetime'],
    }

    @expose('/edit/', methods=('GET', 'POST'))
    def edit_view(self):
        protocolo_id = request.args['id']
        protocolo = Protocolo.query.get_or_404(protocolo_id)
        poste_id = protocolo.item_manutencao.poste_id
        self._template_args['poste_id'] = poste_id
        return super(ProtocoloView, self).edit_view()


class ItemManutencaoView(_ModelView):
    model = ItemManutencao
    name = u'Itens Manutenção'
    category = 'Protocolos'

    column_labels = {
        'criacao': u'Criação'
    }

    def is_accessible(self):
        return current_user.has_role(u'admin')


class OrdemServicoView(_ModelView):
    def __init__(self, *args, **kwargs):
        super(OrdemServicoView, self).__init__(*args, **kwargs)
        self.itens_manutencao_adicionar = None

    def item_manutencao_query(self):
        return ItemManutencao.query.join(Poste).filter(ItemManutencao.status == 'aberto')  # NOQA

    model = OrdemServico
    name = u'Ordem de Serviço'
    category = 'Protocolos'

    form_args = {
        u'criacao': form_args_formats[u'datetime'],
    }

    form_widget_args = {
        u'criacao': form_widget_formats[u'datetime'],
    }

    details_template = 'admin/model/details_os.html'
    form_excluded_columns = ['itens_manutencao']
    column_labels = {
        'criacao': u'Criação',
        'id': u'Número',
    }
    column_display_pk = True
    column_default_sort = ('id', True)
    can_edit = False
    can_view_details = True
    can_delete = False

    column_formatters = {
        'status': lambda v, c, model, n: model.status_map[model.status]
    }

    _urbam_accessible = ['mostrar_pdf', 'enviar_para_servico', 'details_view',
                         'atualizar_item_manutencao']

    column_filters = ['criacao']

    def is_accessible(self):
        if not current_user.has_role('urbam', invert_for_admin=True):
            return super(OrdemServicoView, self).is_accessible()
        for endpoint in self._urbam_accessible:
            if request.endpoint == 'ordemservico.{}'.format(endpoint):
                return True

    def is_visible(self):
        if current_user.has_role('urbam', invert_for_admin=True):
            return False
        return super(OrdemServicoView, self).is_accessible()

    def on_model_change(self, form, ordem_servico, is_created):
        if is_created:
            itens_manutencao = self.itens_manutencao_adicionar
            self.itens_manutencao_adicionar = None
            for item_manutencao in itens_manutencao:
                item_manutencao.status = 'em_servico'
                assoc = Servico()
                assoc.ordem_servico = ordem_servico
                assoc.item_manutencao = item_manutencao
                db.session.add(assoc)
            db.session.commit()

    @expose('/atualizar_item_manutencao/<ordem_servico_id>', methods=['POST'])
    def atualizar_item_manutencao(self, ordem_servico_id):
        def _str_to_bool(value):
            return value == u'true'

        form = request.form
        ordem_servico = OrdemServico.query.get_or_404(ordem_servico_id)
        servico_id = form['servico_id']
        servico = Servico.query.get_or_404(servico_id)
        if ordem_servico.em_servico:
            feito = form['servico_realizado_{}'.format(servico.id)]
            feito = _str_to_bool(feito)
            servico.feito = feito
            servico.resolucao = datetime.now()
            if feito:
                equipamento_keys = [key for key in form.keys() if u'equipamento' in key]
                equipamento_id_quantidade = {}
                for key in equipamento_keys:
                    _, equipamento_id = key.split(u'_')
                    equipamento_id_quantidade[int(equipamento_id)] = int(form[key])
                equipamentos = Equipamento.query \
                    .filter(Equipamento.id.in_(equipamento_id_quantidade.keys()))
                if all(quantidade == 0 for quantidade in equipamento_id_quantidade.values()):
                    flash(u'Serviço deve usar pelo menos um equipamento', u'error')
                    return redirect(request.referrer)
                for equipamento in equipamentos:
                    quantidade = equipamento_id_quantidade[equipamento.id]
                    if quantidade:
                        td = timedelta(days=equipamento.preco_atual.garantia)
                        poste = servico.item_manutencao.poste
                        servicos_poste_equipamento = Servico.query.join(Material) \
                            .join(Equipamento).join(ItemManutencao) \
                            .filter(Servico.id != servico.id, ItemManutencao.poste == poste,
                                    ItemManutencao.fechado, Equipamento.id == equipamento.id) \
                            .order_by(Servico.resolucao.asc()).all()
                        em_garantia = False
                        for _servico in servicos_poste_equipamento:
                            em_garantia = _servico.resolucao + td > datetime.now()
                        material = Material(equipamento=equipamento, servico=servico,
                                            em_garantia=em_garantia, quantidade=quantidade)
                        db.session.add(material)
            else:
                servico.obs_urbam = form[u'comentario_nao_realizacao']
            if all(servico.feito is not None for servico in ordem_servico.servicos):
                ordem_servico.status = 'feita'
        elif ordem_servico.feita:
            fechar = _str_to_bool(form[u'fechar'])
            servico.confirmado = fechar
            if fechar:
                servico.obs_secretaria = form[u'comentario_fechamento']
            if all(servico.confirmado for servico in ordem_servico.servicos):
                ordem_servico.status = 'confirmada'
        else:
            flash(u'Ordem de serviço em status inválido', u'error')
        db.session.commit()
        return redirect(request.referrer)

    @expose('/new/', methods=('GET', 'POST'))
    def create_view(self):
        if request.method == 'POST':
            postes_id = request.form.getlist('postes')
            if not postes_id:
                abort(400)
            query = self.item_manutencao_query().filter(Poste.id.in_(postes_id))
            count = query.count()
            # if count <= 0 or count > 50:
            if count <= 0:
                abort(400)
            self.itens_manutencao_adicionar = query.all()
        return super(OrdemServicoView, self).create_view()

    def _equipamento_query(self):
        return Equipamento.query.filter(Equipamento.ativo == True)  # NOQA

    @expose('/details/', methods=('GET', 'POST'))
    def details_view(self):
        self._template_args['equipamentos'] = self._equipamento_query()
        self._template_args['os_status_map'] = OrdemServico.status_map
        return super(OrdemServicoView, self).details_view()

    def _gerar_pdf(self, template):
        fd, temp_path = mkstemp()
        with open(temp_path, 'wb') as pdf_file:
            pisa.CreatePDF(StringIO(template), pdf_file)
        os.close(fd)
        return temp_path

    @expose('/pdf/<ordem_servico_id>')
    def mostrar_pdf(self, ordem_servico_id):
        ordem_servico = self.model.query.get_or_404(ordem_servico_id)
        template = self.render('pdf/ordem_servico.html', ordem_servico=ordem_servico,
                               equipamentos=self._equipamento_query())
        if not request.args.get('render_html'):
            pdf_path = self._gerar_pdf(template)
            return send_file(pdf_path,  as_attachment=True, mimetype='application/pdf',
                             attachment_filename=
                             u'Ordem de servico #{}.pdf'.format(ordem_servico.id))
        else:
            return template

    @expose('/servico/<ordem_servico_id>', methods=['POST'])
    def enviar_para_servico(self, ordem_servico_id):
        ordem_servico = self.model.query.get_or_404(ordem_servico_id)
        if ordem_servico.nova:
            ordem_servico.status = 'em_servico'
            db.session.commit()
        return redirect(request.referrer)

    @expose('/confirmar/<ordem_servico_id>', methods=['POST'])
    def confirmar(self, ordem_servico_id):
        ordem_servico = self.model.query.get_or_404(ordem_servico_id)
        if ordem_servico.feita:
            ordem_servico.status = 'confirmada'
            db.session.commit()
        return redirect(request.referrer)

    @expose(u'/relatorio')
    def relatorio(self):
        ids = request.args.getlist(u'ids')
        if not ids:
            abort(400)
        ordens_servico = self.model.query.filter(self.model.id.in_(ids), OrdemServico.confirmada) \
            .order_by(self.model.id.desc())
        total_geral = sum(ordem_servico.custo for ordem_servico in ordens_servico)
        return self.render(u'admin/relatorios/ordem_servico.html', ordens_servico=ordens_servico,
                           total_geral=total_geral)

    @action(u'relatorio', u'Relatório')
    def gerar_relatorio(self, ids):
        return redirect(url_for('.relatorio', ids=ids))


class UserView(_UserModelsView):
    model = User
    name = u'Usuários'
    category = u'Usuários'

    column_labels = {
        'password': u'Senha',
        'active': u'Ativo',
    }


class RoleView(_UserModelsView):
    model = Role
    name = u'Role'
    category = u'Usuários'

    column_labels = {
        'name': u'Nome',
        'description': u'Descrição',
    }

    form_excluded_columns = ['users']


class PlanilhaUploadView(SecretariaAcessibleMixin, FileAdmin):
    can_delete_dirs = False
    can_mkdir = False
    can_rename = False
    can_delete = False

    protocolo_sheet_name = u'SituacaoProtocolo 1 '

    allowed_extensions = [u'xlsx']

    def _get_protocolos_from_book(self, workbook):
        sheet = workbook.get_sheet_by_name(self.protocolo_sheet_name)
        protocolos = []
        for row in sheet.iter_rows():
            value = row[0].value
            if not value or (isinstance(value, basestring) and
                            (u'Período' in value or
                             u'RELATÓRIO' in value or
                             u'Protocolo' in value)):
                continue
            cod_protocolo = row[0].value
            logradouro_parts = row[4].value.rsplit(u',', 1)
            logradouro, numero = logradouro_parts[0], int(logradouro_parts[1])
            protocolos.append({
                u'cod_protocolo': cod_protocolo,
                u'criacao': row[1].value,
                u'logradouro': logradouro,
                u'numero': numero,
                u'bairro': row[5].value,
            })
        return protocolos

    @action(u'importar_protocolos', u'Importar Protocolos')
    def importar_protocolos(self, filenames):
        return redirect(url_for('.grade_protocolos', filenames=filenames))

    @expose(u'/grade_protocolos/')
    def grade_protocolos(self):
        filenames = request.args.getlist(u'filenames')
        protocolos = []
        for filename in filenames:
            filename = os.path.join(self.base_path, filename)
            workbook = load_workbook(filename=filename)
            sheet_protocolos = self._get_protocolos_from_book(workbook)
            protocolos.extend(sheet_protocolos)
        # protocolos = protocolos[:10]
        logradouros = Logradouro.query.all()
        contador_status = self._processa_protocolos(protocolos, logradouros)
        return self.render(u'admin/model/grade_protocolos.html', protocolos=protocolos,
                           logradouros=logradouros, contador_status=contador_status)

    def _processa_protocolos(self, protocolos, logradouros):
        logradouros_cache = {logradouro.logradouro: logradouro for logradouro in logradouros}
        contador_status = Counter(['ok', 'poste', 'logradouro'])
        for protocolo in protocolos:
            result = self._similarity_query(protocolo[u'logradouro'])
            row = result.first()
            protocolo[u'id'] = None
            if row:
                logradouro = logradouros_cache[row[u'logradouro']]
                similaridade = row[u'similaridade']
                protocolo[u'similaridade'] = similaridade
                protocolo[u'logradouro_ci'] = logradouro
                protocolo[u'bairro_ci'] = logradouro.bairro
                poste_q = Poste.query.filter_by(logradouro=logradouro)
                poste = poste_q.filter_by(numero=protocolo[u'numero']).first()
                if not poste:
                    if protocolo[u'numero'] == 0:
                        poste = Poste(logradouro=logradouro, numero=0)
                        db.session.add(poste)
                    else:
                        protocolo[u'erro_tipo'] = 'poste'
                        protocolo[u'postes'] = poste_q
                        contador_status['poste'] += 1
                _protocolo = None
                if poste:
                    protocolo[u'poste_id'] = poste.id
                    _protocolo = Protocolo(cod_protocolo=protocolo[u'cod_protocolo'],
                                           criacao=protocolo[u'criacao'], poste=poste)
                    db.session.add(_protocolo)
                    contador_status['ok'] += 1
                db.session.commit()
                if _protocolo:
                    protocolo[u'id'] = _protocolo.id
            else:
                contador_status['logradouro'] += 1
                protocolo[u'erro_tipo'] = u'logradouro_nao_encontrado'
        return contador_status

    def _similarity_query(self, logradouro):
        query = text(u'SELECT (similarity(logradouro.logradouro, :logradouro)) as similaridade,'
                     ' logradouro.logradouro FROM logradouro WHERE logradouro.logradouro %'
                     ' :logradouro ORDER BY similaridade DESC LIMIT 1;')
        params = {u'logradouro': logradouro}
        return db.session.execute(query, params)


def init_app(app):
    config = {
        'url': '/',
    }
    imv = ItemManutencaoView()
    file_upload_path = os.path.join(app.instance_path, u'storage')
    if not os.path.isdir(file_upload_path):
        os.makedirs(file_upload_path)
    views = [
        # endereco
        RegiaoView(),
        BairroView(),
        LogradouroView(),
        # protocolos
        PlanilhaUploadView(base_path=file_upload_path, name=u'Upload de planilha de protocolos',
                           category=u'Protocolos'),
        ProtocoloView(),
        PosteView(imv),
        imv,  # Retirar em prod
        OrdemServicoView(),
        # equipamentos
        EquipamentoView(),
        PrecoEquipamentoView(),
        # usuarios
        UserView(),
        RoleView(),

    ]
    index = IndexView(name='Principal', **config)
    admin = Admin(app, template_mode='bootstrap3', index_view=index, name='Cidade Iluminada',
                  **config)
    for view in views:
        admin.add_view(view)
